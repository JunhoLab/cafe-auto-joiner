from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from typing import List

import openpyxl

STATUS_PENDING = "대기"
STATUS_SUCCESS = "카페가입완료"
STATUS_APPROVAL_PENDING = "카페가입대기"
STATUS_FAILED = "실패"
STATUS_CAPTCHA = "캡차필요"


@dataclass
class AccountRow:
    row_index: int
    naver_id: str
    naver_pw: str
    cafe_url: str
    nickname: str
    spare_nickname: str
    answers: List[str] = field(default_factory=list)
    status: str = STATUS_PENDING

    @property
    def effective_answers(self) -> List[str]:
        return [a for a in self.answers if a and a.strip()]


def load_excel(path: str) -> List[AccountRow]:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    status_col = _find_column(ws, "status", fallback=11)
    rows: List[AccountRow] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        cells = list(row) + [None] * max(0, status_col - len(row))
        if not cells[0]:
            continue
        rows.append(AccountRow(
            row_index=row_idx,
            naver_id=str(cells[0] or "").strip(),
            naver_pw=str(cells[1] or "").strip(),
            cafe_url=str(cells[2] or "").strip(),
            nickname=str(cells[3] or "").strip(),
            spare_nickname=str(cells[4] or "").strip(),
            answers=[str(cells[i] or "").strip() for i in range(5, 10)],
            status=str(cells[status_col - 1] or STATUS_PENDING).strip(),
        ))
    return rows


def update_status(path: str, row_index: int, status: str) -> None:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    status_col = _find_column(ws, "status", fallback=11)
    if status == STATUS_SUCCESS:
        stamp = datetime.now().strftime("%Y-%m-%d %H:%M")
        ws.cell(row=row_index, column=status_col, value=f"카페가입완료 ({stamp})")
    elif status == STATUS_APPROVAL_PENDING:
        stamp = datetime.now().strftime("%Y-%m-%d %H:%M")
        ws.cell(row=row_index, column=status_col, value=f"카페가입대기 ({stamp})")
    else:
        ws.cell(row=row_index, column=status_col, value=status)
    wb.save(path)


def _find_column(ws, header_name: str, fallback: int | None = None) -> int | None:
    target = header_name.strip().lower()
    for col_idx, value in enumerate(ws.iter_rows(min_row=1, max_row=1, values_only=True).__next__(), start=1):
        if str(value or "").strip().lower() == target:
            return col_idx
    return fallback
